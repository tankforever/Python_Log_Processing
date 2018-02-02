# -*- coding: utf-8 -*-
"""
Created on Thu Jan 25 16:04:19 2018

@author: MTK14660
"""

import os
import mace
import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, colors
##################################################################
# 读取calibration table, 只保留标记为y的item    
def read_csv_get_todo_item(cwd,csv_file_name):
    with open(os.path.join(current_work_dir, csv_file_name),'r') as csv_r:
        reader = csv.DictReader(csv_r)
        for rr in reader:
                rlf_time=rr["RLF time"]
    
    return int(rlf_time)
##################################################################
    

##################################################################
# 查找以'MDLog1'为前缀的文件夹
def find_MDLog1(root,search_item):
    
    walk_res=[]
    for root,dirs,files in os.walk(root):
        walk_res.append((root,dirs,files))
    
    all_MDlog1=[]
    for w in walk_res:
        root=w[0]         #root
        basename=os.path.basename(root)
        if search_item in basename:
            #搜索.elg or .muxz
            files = w[2]  #files
            for f in files:
                if '.muxz' in f or '.elg' in f:
                    all_MDlog1.append(w)
                    break
    del walk_res
    
    return all_MDlog1
##################################################################
    

##################################################################
# 找到\check_calibration文件夹下,包含.elg or .muxz文件的子文件夹
def get_root_dirs_path(current_work_dir,all_MDLog1):
    root_dirs_path=[]
    cwd_split_len=len(current_work_dir.split('\\'))
    for i in all_MDLog1:
        temp1=i[0]           #root
        temp2=temp1.split('\\')
        temp3=temp2[0:cwd_split_len+1]
        temp4="\\".join(temp3)
        root_dirs_path.append(temp4)
    root_dirs_path={}.fromkeys(root_dirs_path).keys()   #删除重复元素 
    
    return root_dirs_path
##################################################################

 
################################################################## 
#将root_dirs分类，有相同root的all_MDLog1放在一起
# {root_dir1:[all_MDLog1[0],all_MDLog1[1]], root_dir2:[all_MDLog1[3],all_MDLog1[4]]}    
def classify_same_root(root_dirs_path,all_MDLog1):
    root_dirs_dict=dict()
    for rdp in root_dirs_path:
        temp1=[]
        for aM in all_MDLog1:
            temp2=aM[0].split('\\')    #root
            if os.path.basename(rdp) in temp2:
                temp1.append(aM)
        root_dirs_dict[rdp]=temp1
    
    return root_dirs_dict
##################################################################

##################################################################
# 提取出'MDLog1_2010_0101_081936_data.muxz'中的20100101081936
def parsing_filename(file_name):
    temp1=file_name.split(".")[0]
    temp2=temp1.split("_")[1:4]   # 获取['2017','1201','180008']
    temp3="".join(temp2)
    temp4=int(temp3)
    
    return temp4
##################################################################
    

##################################################################    
#获取全部.muxz or .elg files
#{file1:(root,muxz_file_path,muxz/elg_flag,EDB_file),file2:(),file3:()}
def get_all_muxz_elg_files(root_path,list_root_path_file,muxz_ext,elg_ext):
    muxz_elg_flag=0   # 1:muxz,2:elg
    all_muxz_elg={}
    for r_f_f in list_root_path_file[root_path]:
        root = r_f_f[0]
        files=r_f_f[2]
        
        #先检查是否存在.muxz
        for files_tmp in files:
            if muxz_ext in files_tmp:
                muxz_elg_flag = 1
                break
        #若不存在.muxz, 再检查.elg
        if 0 == muxz_elg_flag:
            for files_tmp in files:
                if elg_ext in files_tmp:
                    muxz_elg_flag = 2
                    break
      
        if 0 == muxz_elg_flag:
            raise ValueError(".muxz & .elg files DO NOT exist!")
        
        if 1 == muxz_elg_flag:
             #得到EDB文件
            EDB_file=[content for content in files if "EDB" == content[-3:] if "MDDB_" == content[0:5]]
            EDB_file_path=os.path.join(root,EDB_file[0])
            
            #得到全部.muxz文件
            for files_tmp in files:
                if muxz_ext in files_tmp:
                    time_in_filename=parsing_filename(files_tmp)
                    if time_in_filename in all_muxz_elg.keys():
                        raise ValueError("repetitive .muxz file name")
                    else:
                        all_muxz_elg[time_in_filename]=(root,files_tmp,muxz_ext,EDB_file_path)
        elif 2 == muxz_elg_flag:
            EDB_file_path=None
            #得到全部.elg文件
            for files_tmp in files:
                if elg_ext in files_tmp:
                    time_in_filename=parsing_filename(files_tmp)
                    all_muxz_elg[time_in_filename]=(root,files_tmp,elg_ext,EDB_file_path)
        else:
            raise ValueError("Wrong muxz_elg_flag value!")
        return all_muxz_elg
##################################################################
   
##################################################################
#二分查找
#def search_item(start,end,search_index,itemset):
#    middle=(start+end)/2
#    tmp_idx=itemset[middle].device_time
#    
#    if  tmp_idx == search_index:
#        return middle
#    elif tmp_idx < search_index:
#        search_item(middle,end,search_index,itemset)
#    else:
#        search_item(start,middle,search_index,itemset)
##################################################################
        
    
##################################################################
# 解析.muxz or .elg log
def parsing_log(root_path,list_root_path_file):
    
    muxz_ext=".muxz"
    elg_ext=".elg"
    #获取全部.muxz or .elg files
    #{文件名:(文件属性:根文件path,muxz/elg文件名,.muxz/.elg后缀,EDB_file_path)}
    #{file1:(root,muxz_file_name,muxz/elg_flag,EDB_file),file2:(),file3:()}
    all_muxz_elg=get_all_muxz_elg_files(root_path,list_root_path_file,muxz_ext,elg_ext)
    
    #对.muxz or .elg file按照时间顺序排序进行排序
    sort_muxz_elg_keys=all_muxz_elg.keys()
    sort_muxz_elg_keys.sort() 
    
    cc_cell_info      = {0:[], 1:[], 2:[], 3:[], 4:[]}
    all_meas_rssi_rscp = []
    all_meas_snr       = []
    all_meas_md        = []
    all_meas_sd        = []
    mixed_cc_cell      = []
    #调用mace模块开始分析log
    for sort_key_idx, sort_key in enumerate(sort_muxz_elg_keys):
        if all_muxz_elg[sort_key][2] == muxz_ext:  #.muxz后缀
            #                                                 root_path            file_name                          
            logfile = mace.open_log_file(os.path.join(all_muxz_elg[sort_key][0],all_muxz_elg[sort_key][1]), database=all_muxz_elg[sort_key][3])
        elif all_muxz_elg[sort_key][2] == elg_ext: #.elg后缀
            logfile = mace.open_log_file(os.path.join(all_muxz_elg[sort_key][0],all_muxz_elg[sort_key][1]))
        
        ######### search: cell info #########
        itemset = mace.create_itemset(logfile)
        cell_info_ps ='GP1_MEAS_SERVING_MEAS_RESULTS'
        itemset.subscribe_ps(cell_info_ps)
        
        len_itemset=len(itemset)
        if len_itemset == 0:
            raise ValueError("{} do not exist!".format(cell_info_ps))
        
        #根据cc进行分类:cc0,cc1,cc2,cc3
        for i in itemset:
            cc_idx = int(i[0])
            #                            PCI      FREQ
            i_tuple = [i.device_time, int(i[1]), int(i[2])]
            cc_cell_info[cc_idx].append(i_tuple)
            mixed_cc_cell.append([i.device_time ,cc_idx])
        
        # 以'GP1_MEAS_SERVING_MEAS_RESULTS'消息的第一项为起始点
        set_begin_time = itemset[0].device_time
        
        #########output: RSSI & RSRP #########
        itemset = mace.create_itemset(logfile)
        meas_rssi_rscp_ps = 'GP1_MEAS_SCM_ONE_SHOT_RESULT_4'
        itemset.subscribe_ps(meas_rssi_rscp_ps)
        # 从一个文件中将'GP1_MEAS_SERVING_MEAS_RESULTS'消息的第一项为起始点
        if 0 == sort_key_idx:
            itemset.set_time_range(set_begin_time)
        
        len_itemset=len(itemset)
        if len_itemset == 0:
            raise ValueError("{} do not exist!".format(meas_rssi_rscp_ps))        
        for i in itemset:
            #                         rsrp_rx0  rsrp_rx1    rssi_rx0   rssi_rx1   cc_flag
            i_tuple = [i.device_time, int(i[2]), int(i[3]), int(i[8]), int(i[9]), None]
            all_meas_rssi_rscp.append(i_tuple)
            
        #########output: snr #########
        itemset = mace.create_itemset(logfile)
        meas_snr_l1 = 'EL1_CH_RX_QUAL_RPT_OS_SNR'
        itemset.subscribe_l1(meas_snr_l1)
        # 从一个文件中将'GP1_MEAS_SERVING_MEAS_RESULTS'消息的第一项为起始点
        if 0 == sort_key_idx:
            itemset.set_time_range(set_begin_time)
        
        len_itemset=len(itemset)
        if len_itemset == 0:
            raise ValueError("{} do not exist!".format(meas_snr_l1))        
        for i in itemset:
            #                          snr_rx0     snr_rx1  cc_flag
            i_tuple = [i.device_time, int(i[1]), int(i[2]), int(i[0])]
            all_meas_snr.append(i_tuple)
          
        #########output: md #########
        itemset = mace.create_itemset(logfile)
        meas_md_l1 = 'EL1D_TRC_RX_DRPT_INNER_MD'
        itemset.subscribe_l1(meas_md_l1)
        # 从一个文件中将'GP1_MEAS_SERVING_MEAS_RESULTS'消息的第一项为起始点
        if 0 == sort_key_idx:
            itemset.set_time_range(set_begin_time)
        
        len_itemset=len(itemset)
        if len_itemset == 0:
            raise ValueError("{} do not exist!".format(meas_md_l1))        
        for i in itemset:
            #                          md_idx    cc_flag
            i_tuple = [i.device_time, int(i[1]), int(i[0])]
            all_meas_md.append(i_tuple)
        
        #########output: sd #########
        itemset = mace.create_itemset(logfile)
        meas_sd_l1 = 'EL1D_TRC_RX_FWS_RPT_COMB_1'
        itemset.subscribe_l1(meas_sd_l1)
        # 从一个文件中将'GP1_MEAS_SERVING_MEAS_RESULTS'消息的第一项为起始点
        if 0 == sort_key_idx:
            itemset.set_time_range(set_begin_time)
        
        len_itemset=len(itemset)
        if len_itemset == 0:
            raise ValueError("{} do not exist!".format(meas_sd_l1))        
        for i in itemset:
            #                          sd_idx     cc_flag
            i_tuple = [i.device_time, int(i[10]), int(i[0])]
            all_meas_sd.append(i_tuple)

    ########## 对RSSI & RSRP添加所属cc的标签 ##########
    cell_idx = 0
    rsrp_idx = 0
    len_mixed_cc_cell = len(mixed_cc_cell)
    len_rsrp          = len(all_meas_rssi_rscp)
#    for idx in range(len(all_meas_rssi_rscp)):
    while 1:
        cc_start = mixed_cc_cell[cell_idx    ][0]   #current time
        cc_end   = mixed_cc_cell[cell_idx + 1][0]   #next time
        cc_flag  = mixed_cc_cell[cell_idx    ][1]   #current cc flag

        rscp_time = all_meas_rssi_rscp[rsrp_idx][0]
        if rscp_time < cc_start:
            rsrp_idx = rsrp_idx + 1
        elif rscp_time >= cc_start and rscp_time < cc_end:
            all_meas_rssi_rscp[rsrp_idx][5] = cc_flag
            rsrp_idx = rsrp_idx + 1
            if rsrp_idx >= len_rsrp:   # 当rsrp_idx == len_rsrp - 1,意味着到达最后一项
                break
        else:
            cell_idx = cell_idx + 1
            if cell_idx >= len_mixed_cc_cell - 1:    #最后一个cell
                cc_flag = mixed_cc_cell[cell_idx][1]
                for rest_idx in range(rsrp_idx, len_rsrp):  # 剩余没有被标记的rscp用最后一个cell的cc flag来标记
                    all_meas_rssi_rscp[rest_idx][5] = cc_flag
                break
        
    ###### 针对不同的cc,对cell_info进行分段:相同的pci&&freq为一块 #######
    cc_cell_info_block = dict()
    for cc_idx in cc_cell_info.keys():
        
        each_cc_cell = cc_cell_info[cc_idx]
        if 0 == len(each_cc_cell):
            continue
        
        block_idx  = 1
        start_time = each_cc_cell[0][0]
        pci        = each_cc_cell[0][1]
        freq       = each_cc_cell[0][2]
        
        meas_data = {"RSRP_RX0":[],"RSRP_RX1":[],"RSSI_RX0":[],"RSSI_RX1":[],"RSRQ_RX0":[],"RSRQ_RX1":[],"SNR_RX0":[],"SNR_RX1":[],"MD_IDX":[],"SD_IDX":[]}
        cell_info_block={block_idx:{"PCI":pci,"FREQ":freq,"START_TIME":start_time,"END_TIME":None,"MEAS_DATA":meas_data}}
        
        for c_i_tmp in each_cc_cell:
            
            dev_time = c_i_tmp[0]
            pci_tmp  = c_i_tmp[1]
            freq_tmp = c_i_tmp[2]
            
            #上一块最后一次的时间是下一块的开始时间
            cell_info_block[block_idx]["END_TIME"] = dev_time
            if pci_tmp != pci or freq_tmp != freq:
                pci, freq=pci_tmp,freq_tmp
                block_idx = block_idx + 1
                #重新创建meas_data,否则cell_info_block中每个item中的meas_data指向同一个地址
                #会导致的结果是，写一个block中的meas_data，其他block的meas_data会同时写入相同的值
                meas_data = {"RSRP_RX0":[],"RSRP_RX1":[],"RSSI_RX0":[],"RSSI_RX1":[],"RSRQ_RX0":[],"RSRQ_RX1":[],"SNR_RX0":[],"SNR_RX1":[],"MD_IDX":[],"SD_IDX":[]}
                cell_info_block[block_idx] = {"PCI":pci,"FREQ":freq,"START_TIME":dev_time,"END_TIME":None,"MEAS_DATA":meas_data}
        #最后一块时间段的end time是最后一个muxz文件的endtime
        cell_info_block[block_idx]["END_TIME"] = logfile.end_time.device_time     
        
        cc_cell_info_block[cc_idx] = cell_info_block
        
    
    ########## 填充RSRP & RSSI ###############
    for cc_idx in cc_cell_info_block.keys():
        cell_info_block_tmp = cc_cell_info_block[cc_idx]
        
        block_i = 1
        for m_r_r_tmp in all_meas_rssi_rscp:
            if cc_idx == m_r_r_tmp[5]:   #c_flag  
                device_time = m_r_r_tmp[0]
                if device_time < cell_info_block_tmp[block_i]["START_TIME"]:
                    pass
                #上一块最后一次的时间是下一块的开始时间，消息的时间需要大于等于start time && 小于（不能小于等于）end time
                elif device_time >= cell_info_block_tmp[block_i]["START_TIME"] and device_time < cell_info_block_tmp[block_i]["END_TIME"]:
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["RSRP_RX0"].append(m_r_r_tmp[1])
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["RSRP_RX1"].append(m_r_r_tmp[2])
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["RSSI_RX0"].append(m_r_r_tmp[3])
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["RSSI_RX1"].append(m_r_r_tmp[4])
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["RSRQ_RX0"].append(m_r_r_tmp[1] - m_r_r_tmp[3])
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["RSRQ_RX1"].append(m_r_r_tmp[2] - m_r_r_tmp[4])
                                    
                else:
                    block_i = block_i + 1
                    if block_i > len(cell_info_block_tmp):
                        break
    
    ########## 填充snr ##############
    for cc_idx in cc_cell_info_block.keys():
        cell_info_block_tmp = cc_cell_info_block[cc_idx]
        
        block_i = 1
        for meas_tmp in all_meas_snr:
            if cc_idx == meas_tmp[3]:
                device_time = meas_tmp[0]
                if device_time < cell_info_block[block_i]["START_TIME"]:
                    pass
                elif device_time >= cell_info_block_tmp[block_i]["START_TIME"] and device_time < cell_info_block_tmp[block_i]["END_TIME"]:
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["SNR_RX0"].append(meas_tmp[1])
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["SNR_RX1"].append(meas_tmp[2])
                                    
                else:
                   block_i = block_i + 1
                   if block_i > len(cell_info_block_tmp):
                       break

    ########## 填充MD ###############
    for cc_idx in cc_cell_info_block.keys():
        cell_info_block_tmp = cc_cell_info_block[cc_idx]

        block_i = 1
        for meas_tmp in all_meas_md:
            if cc_idx == meas_tmp[2]:
                device_time = meas_tmp[0]
                if device_time < cell_info_block_tmp[block_i]["START_TIME"]:
                    pass
                elif device_time >= cell_info_block_tmp[block_i]["START_TIME"] and device_time < cell_info_block_tmp[block_i]["END_TIME"]:
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["MD_IDX"].append(meas_tmp[1])
                    
                else:
                   block_i = block_i + 1
                   if block_i > len(cell_info_block_tmp):
                       break

    ########## 填充SD ############### 
    for cc_idx in cc_cell_info_block.keys():
        cell_info_block_tmp = cc_cell_info_block[cc_idx]
     
        block_i = 1
        for meas_tmp in all_meas_sd:
            if cc_idx == meas_tmp[2]:
                device_time = meas_tmp[0]
                if device_time < cell_info_block[block_i]["START_TIME"]:
                    pass
                elif device_time >= cell_info_block_tmp[block_i]["START_TIME"] and device_time < cell_info_block_tmp[block_i]["END_TIME"]:
                    cell_info_block_tmp[block_i]["MEAS_DATA"]["SD_IDX"].append(meas_tmp[1])
    
                else:
                   block_i = block_i + 1
                   if block_i > len(cell_info_block_tmp):
                       break

    #########将结果保存到xlsx文件中#######   
    ft = Font(bold=True, color=colors.RED)   # 首行字体加粗

    for cc_idx in cc_cell_info_block.keys():
            
        wb = Workbook()
        ws = wb.active
        
        cell_info_block = cc_cell_info_block[cc_idx]
        
        ## sheet1: Summary
        ws.title = "Summary"     #sheet name
        ws['A1'] = 'CELL_NUMBER'     #column title
        ws['B1'] = 'PCI'         #column title
        ws['C1'] = 'FREQ'        #column title
        ws['D1'] = 'START_TIME'  #column title
        ws['E1'] = 'END_TIME'    #column title
        
        # 首行字体加粗
        ws['A1'].font = ft      
        ws['B1'].font = ft   
        ws['C1'].font = ft
        ws['D1'].font = ft
        ws['E1'].font = ft
        
        # 改变列宽
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12


        for i in range(2, len(cell_info_block) + 2):
            ws['A' + str(i)] =  'cell_' + str(i - 1)
            ws['B' + str(i)] = cell_info_block[i - 1]['PCI']
            ws['C' + str(i)] = cell_info_block[i - 1]['FREQ']
            ws['D' + str(i)] = cell_info_block[i - 1]['START_TIME']
            ws['E' + str(i)] = cell_info_block[i - 1]['END_TIME']
        
        ## sheet2~sheetN:将不同的PCI信息写入不同的sheet在
        for i in range(1, len(cell_info_block) + 1):
            ws_cell = wb.create_sheet("cell_" + str(i))   #create new sheet
            
            # 改变列宽
            ws_cell.column_dimensions['C'].width = 12
            ws_cell.column_dimensions['D'].width = 12

            
            ws_cell['A1'] = 'PCI'
            ws_cell['A1'].font = ft
            ws_cell['A2'] = cell_info_block[i]['PCI']
            
            ws_cell['B1'] = 'FREQ'
            ws_cell['B1'].font = ft
            ws_cell['B2'] = cell_info_block[i]['FREQ']
            
            ws_cell['C1'] = 'START TIME'
            ws_cell['C1'].font = ft
            ws_cell['C2'] = cell_info_block[i]['START_TIME']
            
            ws_cell['D1'] = 'END TIME'
            ws_cell['D1'].font = ft
            ws_cell['D2'] = cell_info_block[i]['END_TIME']
            
            ws_cell['E1'] = 'RSRP_RX0'
            ws_cell['E1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['RSRP_RX0']):
                ws_cell['E' + str(j + 2)] = item
                
            ws_cell['F1'] = 'RSRP_RX1' 
            ws_cell['F1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['RSRP_RX1']):
                ws_cell['F' + str(j + 2)] = item
            
            ws_cell['G1'] = 'RSSI_RX0'
            ws_cell['G1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['RSSI_RX0']):
                ws_cell['G' + str(j + 2)] = item
            
            ws_cell['H1'] = 'RSSI_RX1'
            ws_cell['H1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['RSSI_RX1']):
                ws_cell['H' + str(j + 2)] = item
            
            ws_cell['I1'] = 'RSRQ_RX0'
            ws_cell['I1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['RSRQ_RX0']):
                ws_cell['I' + str(j + 2)] = item
            
            ws_cell['J1'] = 'RSRQ_RX1'
            ws_cell['J1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['RSRQ_RX1']):
                ws_cell['J' + str(j + 2)] = item
            
            ws_cell['K1'] = 'SNR_RX0'
            ws_cell['K1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['SNR_RX0']):
                ws_cell['K' + str(j + 2)] = item
            
            ws_cell['L1'] = 'SNR_RX1'
            ws_cell['L1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['SNR_RX1']):
                ws_cell['L' + str(j + 2)] = item
            
            ws_cell['M1'] = 'MD_IDX'
            ws_cell['M1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['MD_IDX']):
                ws_cell['M' + str(j + 2)] = item
            
            ws_cell['N1'] = 'SD_IDX'
            ws_cell['N1'].font = ft
            for j,item in enumerate(cell_info_block[i]['MEAS_DATA']['SD_IDX']):
                ws_cell['N' + str(j + 2)] = item
            
        #保存文件
        output_path=os.path.join(root_path,"output_cc{0}.xlsx".format(str(cc_idx)))
        wb.save(output_path)
    
    print "log:<{0}>".format(os.path.basename(root_path))
    
##################################################################

#=========================== START MAIN ==========================    
if __name__ == "__main__":
    
    current_work_dir=os.getcwd()
    
    # 读取calibration table, 只保留标记为y的item
#    read_csv_file="check_RLF.csv"
#    rlf_time=read_csv_get_todo_item(current_work_dir, read_csv_file)
    
    # csv writer dict header
    #write_csv_header = ["Log"] + item_todo  
    
    # 找到全部以MDLog1为前缀的文件夹&&其中包含了.elg or .muxz文件
    # 返回数据结构为(root,folders,files)
    search_item='MDLog1'    
    all_MDLog1=find_MDLog1(current_work_dir,search_item)
    
    # 找到\check_calibration文件夹下,包含.elg or .muxz文件的子文件夹
    # 即e.g. \check_calibration\A_folder,\check_calibration\B_folder...
    # 得到A_folder,B_folder...
    root_dirs_path=get_root_dirs_path(current_work_dir, all_MDLog1)
    
    #将root_dirs_path分类，有相同root的all_MDLog1放在一起
    # {root_dir_path1:[all_MDLog1[0],all_MDLog1[1]], root_dir_path2:[all_MDLog1[3],all_MDLog1[4]]}
    root_dirs_dict=classify_same_root(root_dirs_path,all_MDLog1) 
        
    # 处理log
    for root_dir_path in root_dirs_path:
        starttime=datetime.datetime.now()
        parsing_log(root_dir_path,root_dirs_dict)
        endtime=datetime.datetime.now()
        
        print "Executing time: " + str((endtime-starttime).seconds) + "s"
        
#=========================== END MAIN ==========================  